# -*- coding: utf-8 -*-
import pandas as pd
import openpyxl
import streamlit as st
import requests
import matplotlib.pyplot as plt
from fpdf import FPDF
from sklearn.ensemble import RandomForestRegressor
from xgboost import XGBRegressor
from sklearn.linear_model import LinearRegression
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_absolute_error

# Streamlit Web App Title
st.title("AI-Powered Appraisal Adjustment Tool")

# Upload MLS Data File
uploaded_file = st.file_uploader("Upload MLS Data (CSV or Excel)", type=["csv", "xlsx"])

if uploaded_file:
    # Load Data
    file_extension = uploaded_file.name.split(".")[-1]
    if file_extension == "csv":
        df = pd.read_csv(uploaded_file, encoding="utf-8")
    else:
        df = pd.read_excel(uploaded_file)

    # Display Uploaded Data
    st.write("Uploaded MLS Data:")
    st.dataframe(df.head())

    # Ensure Data Cleaning
    df['Total Bedrooms'] = pd.to_numeric(df['Total Bedrooms'], errors='coerce')
    df['Total SqFt.'] = pd.to_numeric(df['Total SqFt.'], errors='coerce')
    df['Total Bathrooms'] = pd.to_numeric(df['Total Bathrooms'], errors='coerce')
    df['Garage Stall'] = pd.to_numeric(df['Garage Stall'], errors='coerce')
    df['Lot Acres'] = pd.to_numeric(df['Lot Acres'], errors='coerce')
    df['Sold Price'] = pd.to_numeric(df['Sold Price'], errors='coerce')
    df_clean = df.dropna()

    # Define Features and Target
    features = ['Total Bedrooms', 'Total SqFt.', 'Total Bathrooms', 'Garage Stall', 'Lot Acres']
    X = df_clean[features]
    y = df_clean['Sold Price']

    # Train-Test Split
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

    # Train Models
    rf_model = RandomForestRegressor(n_estimators=100, random_state=42)
    xgb_model = XGBRegressor(n_estimators=100, learning_rate=0.1, random_state=42)
    lr_model = LinearRegression()

    rf_model.fit(X_train, y_train)
    xgb_model.fit(X_train, y_train)
    lr_model.fit(X_train, y_train)

    # Predict and Average Results
    rf_pred = rf_model.predict(X_test)
    xgb_pred = xgb_model.predict(X_test)
    lr_pred = lr_model.predict(X_test)
    final_pred = (rf_pred + xgb_pred + lr_pred) / 3

    # Calculate Model Performance
    final_mae = mean_absolute_error(y_test, final_pred)
    st.write(f"Model Accuracy (MAE): ${final_mae:,.2f}")

    # Integrate Real-Time Market Data from Zillow API
    st.write("### Real-Time Market Data")
    api_url = "https://api.zillow.com/property-data"
    params = {"zipcode": "83814", "property_type": "single_family"}
    headers = {"Authorization": "Bearer YOUR_ZILLOW_API_KEY"}
    response = requests.get(api_url, params=params, headers=headers)
    if response.status_code == 200:
        market_data = response.json()
        st.write(market_data)
    else:
        st.write("Failed to fetch market data from Zillow API.")

    # Calculate Feature Importance
    feature_importance = dict(zip(features, rf_model.feature_importances_))
    adjustments = {
        "Feature": features,
        "Monetary Adjustment ($)": [feature_importance[feature] * (y.max() - y.min()) for feature in features],
        "Percentage Adjustment (%)": [(feature_importance[feature] * (y.max() - y.min())) / y.median() * 100 for feature in features]
    }
    df_adjustments = pd.DataFrame(adjustments)

    # Display Adjustments
    st.write("### Calculated Appraisal Adjustments:")
    st.dataframe(df_adjustments)

    # Plot Feature Importance
    fig, ax = plt.subplots()
    ax.barh(df_adjustments['Feature'], df_adjustments['Monetary Adjustment ($)'])
    ax.set_xlabel("Monetary Adjustment ($)")
    ax.set_title("Feature Importance in Appraisal Adjustments")
    st.pyplot(fig)

    # Save to Excel
    excel_file_path = "appraisal_adjustments.xlsx"
    df_adjustments.to_excel(excel_file_path, index=False, engine='openpyxl')
    st.download_button(label="Download Adjustments Report (Excel)", data=open(excel_file_path, "rb").read(), file_name=excel_file_path, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # Generate PDF Report
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt="AI-Powered Real Estate Appraisal Report", ln=True, align="C")
    pdf.cell(200, 10, txt=f"Model Accuracy (MAE): ${final_mae:,.2f}", ln=True, align="L")
    pdf.cell(200, 10, txt="Calculated Adjustments:", ln=True, align="L")
    for index, row in df_adjustments.iterrows():
        pdf.cell(200, 10, txt=f"{row['Feature']}: ${row['Monetary Adjustment ($)']:,.2f}", ln=True, align="L")
    pdf_file_path = "appraisal_report.pdf"
    pdf.output(pdf_file_path)
    st.download_button(label="Download Appraisal Report (PDF)", data=open(pdf_file_path, "rb").read(), file_name=pdf_file_path, mime="application/pdf")
